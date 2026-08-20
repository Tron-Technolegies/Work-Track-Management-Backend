from openpyxl import Workbook
from openpyxl.styles import Font
from django.http import HttpResponse


def export_to_excel(
    filename,
    headers=None,
    rows=None,
    sheets=None
):

    workbook = Workbook()

    # =====================================================
    # MULTIPLE SHEETS
    # =====================================================

    if sheets:

        default_sheet = workbook.active

        workbook.remove(
            default_sheet
        )

        for sheet_name, sheet_data in sheets.items():

            worksheet = workbook.create_sheet(
                title=sheet_name[:31]
            )

            sheet_headers = (
                sheet_data.get(
                    "headers",
                    []
                )
            )

            sheet_rows = (
                sheet_data.get(
                    "rows",
                    []
                )
            )

            # Header
            for column, header in enumerate(
                sheet_headers,
                start=1
            ):

                cell = worksheet.cell(
                    row=1,
                    column=column
                )

                cell.value = header

                cell.font = Font(
                    bold=True
                )

            # Data
            for row_index, row in enumerate(
                sheet_rows,
                start=2
            ):

                for column_index, value in enumerate(
                    row,
                    start=1
                ):

                    worksheet.cell(
                        row=row_index,
                        column=column_index,
                        value=value
                    )

            # Column widths
            for column_cells in worksheet.columns:

                max_length = 0

                column_letter = (
                    column_cells[0]
                    .column_letter
                )

                for cell in column_cells:

                    try:

                        length = len(
                            str(cell.value)
                        )

                        max_length = max(
                            max_length,
                            length
                        )

                    except Exception:
                        pass

                worksheet.column_dimensions[
                    column_letter
                ].width = min(
                    max_length + 2,
                    50
                )

    # =====================================================
    # OLD SINGLE-SHEET BEHAVIOUR
    # =====================================================

    else:

        worksheet = workbook.active

        worksheet.title = (
            filename[:31]
        )

        headers = headers or []
        rows = rows or []

        for column, header in enumerate(
            headers,
            start=1
        ):

            cell = worksheet.cell(
                row=1,
                column=column
            )

            cell.value = header

            cell.font = Font(
                bold=True
            )

        for row_index, row in enumerate(
            rows,
            start=2
        ):

            for column_index, value in enumerate(
                row,
                start=1
            ):

                worksheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value
                )

    # =====================================================
    # RESPONSE
    # =====================================================

    response = HttpResponse(
        content_type=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        f'attachment; '
        f'filename="{filename}.xlsx"'
    )

    workbook.save(response)

    return response